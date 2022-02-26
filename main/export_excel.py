from openpyxl import Workbook
from django.http import HttpResponse
from openpyxl.styles import Font,Alignment
from .models import *

def export_wo_xls(request,wo_id):
	my_wo = work_order.objects.all().filter(id=wo_id).first()
	response = HttpResponse(
		content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
	)
	response['Content-Disposition'] = f'attachment; filename={my_wo} Indents.xlsx'
	workbook = Workbook()

	
	# Get active worksheet/tab
	worksheet = workbook.active
	worksheet.title = f'{my_wo} Indents'

	# Define the titles for columns
	col_dict = {
		'Indent ID':'id',
		'Description':'description',
		'Material Type':'material_type',
		'Material Shape':'material_shape',
		'Quantity':'quantity',
		'Weight':'get_weight',
		'Unit value':'value',
		'Tax (in %)':'tax',
		'Tax Value':'tax_amount',
		'Other Expanses':'other_expanses',
		'Discount':'discounted_total',
		'Gross Value':'gross_value',

	}
	
	row_num = 1

	all_indents = indent.objects.all().filter(WO=my_wo)

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(col_dict, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.font = Font(name='Calibri', bold=True, size=12)
		cell.alignment=Alignment(horizontal='left')
		cell.value = column_title


	# Iterate through all movies
	for my_indent in all_indents:
		row_num += 1
		# Define the data for each cell in the row 
		row = []
		for i in col_dict:
			temp = getattr(my_indent, col_dict[i])
			# print(temp)
			if str(type(temp)) == "<class 'method'>":
				row.append(temp())
			else:
				row.append(temp)
				# Assign the data for each cell of the row 
		for col_num, cell_value in enumerate(row, 1):
			cell = worksheet.cell(row=row_num, column=col_num)
			cell.alignment=Alignment(horizontal='left')
			cell.value = cell_value

	workbook.save(response)

	return response

def export_po_xls(request,po_id):
	my_po = purchase_order.objects.all().filter(id=po_id).first()
	response = HttpResponse(
		content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
	)
	response['Content-Disposition'] = f'attachment; filename={my_po.po_number} Indents.xlsx'
	workbook = Workbook()

	
	# Get active worksheet/tab
	worksheet = workbook.active
	worksheet.title = f'{my_po.po_number} Indents'

	# Define the titles for columns
	col_dict = {
		'Indent ID':'id',
		"WO Number ID":'get_wo_number',
		'Description':'description',
		'Material Type':'material_type',
		'Quantity':'quantity',
		'Weight':'get_weight',
		'Unit value':'value',
		'Tax (in %)':'tax',
		'Tax Value':'tax_amount',
		'Other Expanses':'other_expanses',
		'Discount':'discounted_total',
		'Gross Value':'gross_value',
		'Note':'comment',
	}
	
	row_num = 1

	all_indents = indent.objects.all().filter(PO=my_po)

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(col_dict, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.font = Font(name='Calibri', bold=True, size=12)
		cell.alignment=Alignment(horizontal='left')
		cell.value = column_title


	# Iterate through all movies
	for my_indent in all_indents:
		row_num += 1
		# Define the data for each cell in the row 
		row = []
		for i in col_dict:
			temp = getattr(my_indent, col_dict[i])
			# print(temp)
			if str(type(temp)) == "<class 'method'>":
				row.append(temp())
			else:
				row.append(temp)
				# Assign the data for each cell of the row 
		for col_num, cell_value in enumerate(row, 1):
			cell = worksheet.cell(row=row_num, column=col_num)
			cell.alignment=Alignment(horizontal='left')
			cell.value = cell_value

	workbook.save(response)

	return response

